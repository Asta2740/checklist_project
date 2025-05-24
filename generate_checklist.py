


def main():
    print(r""" 

       _               _    _ _     _                              
   ___| |__   ___  ___| | _| (_)___| |_                            
  / __| '_ \ / _ \/ __| |/ / | / __| __|                           
 | (__| | | |  __/ (__|   <| | \__ \ |_                            
  \___|_| |_|\___|\___|_|\_\_|_|___/\__|                           
       __   __                        __   _                       
       \ \ / /__  _   _ ___ ___  ___ / _| | |__   __ _ _ __  _   _ 
        \ V / _ \| | | / __/ __|/ _ \ |_  | '_ \ / _` | '_ \| | | |
         | | (_) | |_| \__ \__ \  __/  _| | | | | (_| | | | | |_| |
         |_|\___/ \__,_|___/___/\___|_|   |_| |_|\__,_|_| |_|\__, |
                                                             |___/ 
                                                                                                                                       
""")
    from src.bluetooth.bluetooth_check import check_bluetooth_scan

    ok, result = check_bluetooth_scan(timeout=8.0)
    if ok:
        print(f"Scan succeeded — found {len(result)} device(s):")
        for addr, name in result:
            print(f"  • {name}  [{addr}]")
        bluetooth_result = f"Scan succeeded — found {len(result)} device(s)"
    else:
        print("Scan failed:", result)
        bluetooth_result = f"Scan failed: {result}"

    from src.keyboard.keyboard_test import prompt_user_gui , run_aqua_key_test
    from src.credentials.serial_number import get_serial_number
    from src.credentials.brand_version import get_brand_and_version
    from src.credentials.department import get_department
    from src.battery.battery_status import get_battery_status
    from src.battery.battery_wear_level import get_battery_wear_level_windows
    from src.webcam.webcam_test import camera_test
    from src.audio.microphone_test import (
        play_audio_test,
        record_microphone_test,
        prompt_user as microphone_prompt_user,
        OUTPUT_FILENAME,
        audio,
        CHUNK
    )
    import wave 
    from src.hard_disk.disk_benchmark import Hard_Disk_check
    from src.hard_disk.disk_info import  get_disk_info , get_all_disk_health_percent
    from src.monitor.monitor_info import get_screen_size
    from src.ram.ram_info import get_ram_info
    from src.cpu.general_cpu import stress_cpu , get_cpu_temps , get_cpu_name


    print("Starting system checks...")

    # Keyboard tests
    keyboard_response = run_aqua_key_test()
    print(keyboard_response)

    # Credentials
    serial = get_serial_number()
    brand_and_version = get_brand_and_version()
    department = get_department()
    print(f"Serial Number: {serial}")
    print(f"Brand and Version: {brand_and_version}")
    print(f"Department: {department}")



    # Webcam test
    webcam_result = camera_test()
    print(webcam_result)

    # Audio tests
    if record_microphone_test():
        try:
            print("Playing back the recording...")
            wf = wave.open(OUTPUT_FILENAME, 'rb')
            stream = audio.open(format=audio.get_format_from_width(wf.getsampwidth()),
                                channels=wf.getnchannels(),
                                rate=wf.getframerate(),
                                output=True)
            data = wf.readframes(CHUNK)
            while data:
                stream.write(data)
                data = wf.readframes(CHUNK)
            stream.stop_stream()
            stream.close()
            print("Playback finished.")
            microphone_test_result = microphone_prompt_user()
        except Exception as e:
            print(f"Error during playback: {e}")
            microphone_test_result = microphone_prompt_user()
    else:
        print("Microphone test failed. Skipping playback.")
        microphone_test_result = microphone_prompt_user()


    
    audio_channel_test_result = play_audio_test()
    print(microphone_test_result)
    print(audio_channel_test_result)


    # Hard disk tests
    health_info = get_all_disk_health_percent()

    all_disks = get_disk_info()
    health_info = list(health_info.items())
    disks_info = []
    i = 0
    for idx, disk in enumerate(all_disks, 1):
        Device, health_percent = health_info[i]
        if "Error reading SMART" in str(health_percent):
            health_percent = "No SMART support"
            disks_info.append(f"Disk {idx}:\nHealth ≈ {health_percent}\nName={disk['name']}\nSize={disk['size_gb']} GB\nType={disk['type']}\n")
        else:
            disks_info.append(f"Disk {idx}:\nHealth ≈ {health_percent} %\nName={disk['name']}\nSize={disk['size_gb']} GB\nType={disk['type']}\n")

        i += 1
    disks_info = "".join(disks_info)
    print(disks_info)

    Disk_result = Hard_Disk_check()
    print(Disk_result)



    # screen size test

    screen_size = get_screen_size()
    print(screen_size)

    ## Ram test
    try:
        get_total_ram_result = get_ram_info()
        print(get_total_ram_result)
    except Exception as e:
        print(f"Error retrieving RAM info: {e}")
        get_total_ram_result= f"Error retrieving RAM info"

    # CPU name test
    cpu_name = get_cpu_name()
    print(f"CPU Name: {cpu_name}")

    # CPU temperature test
    cpu_test_results = stress_cpu(target_usage=90)
    print("CPU stress test completed.")

    # Build the structured string
    cpu_report = (
        f"Average CPU Usage: {cpu_test_results['avg_usage']:.1f}%\n"
        f"Min CPU Usage: {cpu_test_results['min_usage']:.1f}%\n"
        f"Max CPU Usage: {cpu_test_results['max_usage']:.1f}%"
    )

    for core, temps in cpu_test_results['temps'].items():
        cpu_report += (
            f"\n{core}:\n Maximum : {temps['max']:.1f} °C Average : {temps['avg']:.1f} °C"
        )

    print(cpu_report)

    # Battery tests
    Battery_results = get_battery_status()

    Battery_results += get_battery_wear_level_windows()
    print(Battery_results)
    
    print("Battery status and wear level retrieved.")

    # Import required modules from openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import platform
    import os
    from datetime import date , datetime


    wb = Workbook()
    sheet = wb.active
    sheet.title = "Check list 1"


    # Define styles
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="DDDDDD")  # Light grey
    thick = Side(border_style="thick", color="000000")
    header_border = Border(top=thick, bottom=thick, left=thick, right=thick)

    # Define formatting styles
    bold_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))

    # Set title rows with merged cells
    sheet.merge_cells('A1:D1')
    sheet['A1'] = "Generated Check List"
    sheet['A1'].font = header_font
    sheet['A1'].fill = header_fill
    sheet['A1'].border = header_border

    sheet.merge_cells('A2:D2')
    sheet['A2'] = "Check List per Laptop"
    sheet['A2'].font = header_font
    sheet['A2'].fill = header_fill
    sheet['A2'].border = header_border


    # Set input fields
    input_fields = [
        ("Vendor Name:", "Brand:"),
        ("Laptop Serial:", "Asset ID:"),
        ("Time:", "Date:"),
        ("Eng. Name:", "Model:")
    ]
    for i, (label1, label2) in enumerate(input_fields, start=3):
        sheet[f'A{i}'] = label1
        sheet[f'C{i}'] = label2


    sheet['B4'] = serial
    sheet['B4'].font = header_font

    sheet['D6'] = brand_and_version
    sheet['D6'].font = header_font

    sheet['D3'] = department
    sheet['D3'].font = header_font



    Time_now = datetime.now()
    current_time = Time_now.strftime("%H:%M:%S")

    sheet['B5'] = current_time
    sheet['B5'].font = header_font

    today = date.today()
    sheet['D5'] = today.strftime("%d/%m/%Y")
    sheet['D5'].font = header_font
    # Set Hardware Check List
    sheet.merge_cells('B7:D7')
    sheet['A7'] = "Summary Report"
    sheet['A7'].font = header_font

    sheet.merge_cells('A8:D8')
    sheet['A8'] = "Hardware Check List"
    sheet['A8'].font = header_font
    sheet['A8'].fill = header_fill
    sheet['A8'].border = header_border

    # Headers for check list
    headers = ["Item", "Problem Found", "Hardware CMDB", "Note"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border

    # Hardware items
    hardware_items = [
        "CPU","CPU FAN", "RAM", "USB ports working normally", "Screen", "Harddisk", 
        "Body", "Keyboard", "Touchpad", "Charger", "Charging port", "speakers", 
        "Mic", "webcam", "Battery runtime", 
        "Wireless-Bluetooth", "IO"
    ]

    # Data validation options for Hardware CMDB
    cmdb_options = {
        "CPU": "2nd,3rd,4th,5th,6th,7th,8th,9th,10th",
        "RAM": "4GB,8GB,16GB,32GB,64GB",
        "Screen": "13inch,14inch,15inch,16inch,17inch"
    }

    # Add hardware items and data validation
    dv_yes_no = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    sheet.add_data_validation(dv_yes_no)

    Notes = [""] * len(hardware_items)
    ## 0 is CPU
    Notes[0] = cpu_name
    Notes[1] = cpu_report
    ## 2 is CPU FAN

    try:
        Notes[2] = get_total_ram_result
    except Exception as e:
        Notes[2] = get_total_ram_result

    ## 4 is USB ports working normally
    Notes[4] = screen_size



    Notes[5] = str(disks_info)
    Notes[6] = Disk_result

    Notes[7] = keyboard_response

    Notes[11] = audio_channel_test_result
    Notes[12] = microphone_test_result

    Notes[13] = webcam_result
    Notes[14] = Battery_results
    Notes[15] = bluetooth_result

    j = 0
    for i, item in enumerate(hardware_items, start=10):

        # Item name
        sheet[f'A{i}'] = item
        dv_yes_no.add(f'B{i}')


        if j == 0:
            sheet[f'C{i}'] = Notes[j]
            sheet[f'D{i}'] = Notes[j+1]
        elif j == 1:
            j += 1
            continue
        elif j == 2:
            sheet[f'C{i}'] = Notes[j]
        elif j == 4:
            sheet[f'C{i}'] = Notes[j]
        elif j == 5:
            sheet[f'C{i}'] = Notes[j]
            sheet[f'D{i}'] = Notes[j+1]
        elif j == 6:
            j += 1
            continue
        else:
            sheet[f'D{i}'] = Notes[j]
        
        # Data validation for "Problem Found" (Yes/No)

        j += 1

    # Set Software Check List
    software_start_row = 9 + len(hardware_items) + 1  # Row 26
    sheet.merge_cells(f'A{software_start_row}:D{software_start_row}')
    sheet[f'A{software_start_row}'] = "Software Check List"
    sheet[f'A{software_start_row}'].font = header_font
    sheet[f'A{software_start_row}'].fill = header_fill
    sheet[f'A{software_start_row}'].border = header_border

    # Headers for software check list
    for col, header in enumerate(["Item", "Problem Found", "Software CMDB", "Note"], start=1):
        cell = sheet.cell(row=software_start_row + 1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border

    # Software items
    software_items = ["OS", "MS Office", "Chrome", "7-ZIP", "Adobe Reader"]
    for i, item in enumerate(software_items, start=software_start_row + 2):
        sheet[f'A{i}'] = item
        dv_yes_no.add(f'B{i}')  # Yes/No for "Problem Found"

    # Set Spoc fields
    spoc_row = software_start_row + 2 + len(software_items)  # Row 33
    sheet[f'A{spoc_row}'] = "Spoc. Name:"
    sheet[f'C{spoc_row}'] = "Spoc. Signature:"

    # Apply borders to all cells
    for row in range(1, spoc_row + 1):
        for col in ['A', 'B', 'C', 'D']:
            sheet[f'{col}{row}'].border = thin_border

    # Set column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 30

    # Save the workbook with error handling

    if not serial or serial.strip().upper() == "N/A":
        filename = "Generated Check List.xlsx"
    else:
        filename = f"{serial}.xlsx"

    # Get today's date
    today_date = date.today().strftime("%Y-%m-%d")

    # Network paths list
    network_paths = [
##        r"\\network\path\to\save\location1",
    ]

    def try_save_to_network(base_path, filename, date_str, workbook):
        try:
            dated_folder_path = os.path.join(base_path, date_str)
            if not os.path.exists(dated_folder_path):
                os.makedirs(dated_folder_path)
            full_path = os.path.join(dated_folder_path, filename)
            workbook.save(full_path)
            return True, None
        except Exception as e:
            return False, str(e)

    failed_paths = []

    for path in network_paths:
        success, error = try_save_to_network(path, filename, today_date, wb)
        if not success:
            failed_paths.append((path, error))

    if failed_paths:
        print("\nThe following network paths failed to save the workbook:")
        for fp, err in failed_paths:
            print(f"- {fp}: {err}")
        print("\nFalling back to local save...")

        fallback_path = os.getcwd()
        local_dated_folder_path = os.path.join(fallback_path, today_date)
        if not os.path.exists(local_dated_folder_path):
            os.makedirs(local_dated_folder_path)
        local_full_path = os.path.join(local_dated_folder_path, filename)
        try:
            wb.save(local_full_path)
            print(f"Workbook saved successfully at fallback location '{local_full_path}'")
        except Exception as e:
            print(f"Failed to save workbook even at fallback location: {e}")
    else:
        print("\nWorkbook saved successfully on all network paths!")

if __name__ == "__main__":
    main()